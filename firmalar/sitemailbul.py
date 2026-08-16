import pandas as pd
import time
import random
import re

from difflib import SequenceMatcher
from urllib.parse import quote, urlparse

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# =====================================================
# AYARLAR
# =====================================================

MAX_GOOGLE_SONUC = 15

MIN_GOOGLE_PUAN = 70

# Site kimliği için minimum puan
MIN_SITE_PUAN = 70

BEKLEME = (1.5, 2.5)

MAX_ILETISIM_SAYFASI = 5

MAX_BODY = 15000


# =====================================================
# CHROME
# =====================================================

options = Options()

options.debugger_address = "127.0.0.1:9222"

driver = webdriver.Chrome(
    service=Service(
        ChromeDriverManager().install()
    ),
    options=options
)


# =====================================================
# EXCEL
# =====================================================

df = pd.read_excel("firmalar.xlsx")

print("Excel sütunları:")
print(df.columns.tolist())

print("\nToplam firma:", len(df))


# =====================================================
# İSTANBUL İLÇELERİ
# =====================================================

ISTANBUL_ILCELERI = {
    "adalar",
    "arnavutkoy",
    "atasehir",
    "avcilar",
    "bagcilar",
    "bahcelievler",
    "bakirkoy",
    "basaksehir",
    "bayrampasa",
    "besiktas",
    "beykoz",
    "beylikduzu",
    "beyoglu",
    "buyukcekmece",
    "catalca",
    "cekmekoy",
    "esenler",
    "esenyurt",
    "eyupsultan",
    "fatih",
    "gaziosmanpasa",
    "gungoren",
    "kadikoy",
    "kagithane",
    "kartal",
    "kucukcekmece",
    "maltepe",
    "pendik",
    "sancaktepe",
    "sariyer",
    "silivri",
    "sultanbeyli",
    "sultangazi",
    "sisli",
    "sile",
    "tuzla",
    "umraniye",
    "uskudar",
    "zeytinburnu"
}


# =====================================================
# IGNORE
# =====================================================

IGNORE = {

    "google.com",
    "gstatic.com",

    "linkedin.com",
    "facebook.com",
    "instagram.com",
    "youtube.com",
    "twitter.com",
    "x.com",

    "kariyer.net",
    "bulurum.com",
    "find.com.tr",
    "118.com.tr",
    "firmasec.com",
    "firmabulucu",
    "firmaatlas.com",
    "mukellef.info",
    "infobel.com",
    "kompass",
    "europages",
    "emis.com",
    "verif.com",
    "listofcompany.com",
    "tesisat.com.tr",
    "manuzone.com",
    "synevo.com.tr",
    "zavis.ai",
    "sahibinden.com",
    "alibaba.com",
    "amazon.com",
    "trendyol.com",
    "haberler.com",
    "emlakkulisi.com",
    "emlakkulisi",
    "medium.com",
    "eksisozluk.com",

    "ticaretsicil.gov.tr",
    "ito.org.tr",

    ".gov.tr",
    ".edu.tr",
    ".bel.tr",
    ".k12.tr",

    "ihalepro.com",
    "ihalekik.com",
    "ihaleciler.com",
    "ihale.com",
    "kamubilgisistemi.com"
}


# =====================================================
# TÜRKÇE TEMİZLE
# =====================================================

def temizle(text):

    text = str(text).lower()

    text = text.replace("i̇", "i")

    ceviri = str.maketrans(
        "çğıöşü",
        "cgiosu"
    )

    return text.translate(ceviri)


# =====================================================
# URL TEMİZLE
# =====================================================

def domain_adi(url):

    try:

        domain = urlparse(
            url
        ).netloc.lower()

        domain = domain.split(":")[0]

    except:

        domain = str(url).lower()

    domain = domain.replace(
        "www.",
        ""
    )

    uzantilar = [
        ".com.tr",
        ".net.tr",
        ".org.tr",
        ".gen.tr",
        ".com",
        ".net",
        ".org",
        ".co"
    ]

    for uzanti in uzantilar:

        if domain.endswith(uzanti):

            domain = domain[
                :-len(uzanti)
            ]

            break

    return temizle(domain)


# =====================================================
# STOPWORDS
# =====================================================

STOPWORDS = {

    "sanayi",
    "ticaret",
    "limited",
    "ltd",
    "ltdsti",
    "sirketi",
    "anonim",
    "anonimsirketi",
    "as",
    "ve",
    "ithalat",
    "ihracat",
    "hizmet",
    "hizmetleri",
    "teknoloji",
    "teknolojileri",
    "sistem",
    "sistemleri",
    "cozum",
    "cozumleri",
    "makine",
    "makina",
    "muhendislik",
    "mimarlik",
    "insaat",
    "enerji",
    "lojistik",
    "metal",
    "plastik",
    "tekstil",
    "gida",
    "otomotiv",
    "kimya",
    "tasarim",
    "tasarimi",
    "elektrik",
    "elektronik",
    "tic",
    "san",
    "denetim",
    "kalite",
    "guvenlik",
    "cevre",
    "laboratuvar",
    "hastane",
    "geri",
    "donusum",
    "yapi",
    "danismanlik",
    "mobilya",
    "taahhut"
}


# =====================================================
# FİRMA KELİMELERİ
# =====================================================

def firma_kelimeleri(unvan):

    unvan = temizle(unvan)

    kelimeler = re.findall(
        r"[a-z0-9]+",
        unvan
    )

    sonuc = []

    for kelime in kelimeler:

        if len(kelime) <= 1:
            continue

        if kelime.isdigit():
            continue

        if kelime in STOPWORDS:
            continue

        sonuc.append(kelime)

    return sonuc


# =====================================================
# DOMAIN / MARKA BENZERLİĞİ
# =====================================================

def benzerlik(a, b):

    return SequenceMatcher(
        None,
        temizle(a),
        temizle(b)
    ).ratio()


def domain_firma_eslesmesi(unvan, url):

    domain = domain_adi(url)

    kelimeler = firma_kelimeleri(
        unvan
    )

    if not kelimeler:
        return 0

    puan = 0

    for kelime in kelimeler:

        if kelime in domain:

            if len(kelime) >= 8:
                puan += 100

            elif len(kelime) >= 5:
                puan += 70

            else:
                puan += 40

    return puan


# =====================================================
# GOOGLE PUANI
# =====================================================

def google_puani(unvan, url, baslik):

    domain = domain_adi(url)

    baslik = temizle(
        baslik
    )

    kelimeler = firma_kelimeleri(
        unvan
    )

    if not kelimeler:
        return 0

    puan = 0

    domain_eslesen = 0
    title_eslesen = 0

    # -------------------------------------------------
    # DOMAIN
    # -------------------------------------------------

    for kelime in kelimeler:

        if kelime in domain:

            domain_eslesen += 1

            if len(kelime) >= 8:

                puan += 100

            elif len(kelime) >= 5:

                puan += 70

            else:

                puan += 35


    # -------------------------------------------------
    # TITLE
    # -------------------------------------------------

    for kelime in kelimeler:

        if kelime in baslik:

            title_eslesen += 1


    if title_eslesen >= 3:

        puan += 70

    elif title_eslesen == 2:

        puan += 45

    elif title_eslesen == 1:

        puan += 20


    # -------------------------------------------------
    # DOMAIN BENZERLİĞİ
    # -------------------------------------------------

    en_iyi = 0

    for kelime in kelimeler:

        oran = benzerlik(
            kelime,
            domain
        )

        if oran > en_iyi:

            en_iyi = oran

    puan += int(
        en_iyi * 30
    )


    # -------------------------------------------------
    # DOMAIN VE TITLE İLİŞKİSİ YOKSA ELE
    # -------------------------------------------------

    if (
        domain_eslesen == 0
        and title_eslesen == 0
    ):

        return 0


    return puan


# =====================================================
# MAIL BUL
# =====================================================

def mailleri_bul(metin):

    pattern = r"""
        [a-zA-Z0-9._%+\-]+
        @
        [a-zA-Z0-9.\-]+
        \.
        [a-zA-Z]{2,}
    """

    mailler = re.findall(
        pattern,
        metin,
        re.VERBOSE
    )

    sonuc = []

    for mail in mailler:

        mail = mail.lower().strip()

        # Hatalı / saçma uzantılar
        if len(mail) > 100:
            continue

        # Dosya gibi görünen şeyler
        if mail.endswith(
            (".png", ".jpg", ".jpeg", ".gif")
        ):
            continue

        if mail not in sonuc:

            sonuc.append(mail)

    return sonuc


# =====================================================
# MAIL ÖNCELİK
# =====================================================

def mail_puani(mail, site_domain):

    mail = mail.lower().strip()

    try:

        kullanici, mail_domain = mail.split(
            "@",
            1
        )

    except:

        return -1


    puan = 0

    site_domain = temizle(
        site_domain
    )

    mail_domain = temizle(
        mail_domain
    )


    # -------------------------------------------------
    # KURUMSAL DOMAIN
    # -------------------------------------------------

    if site_domain and site_domain in mail_domain:

        puan += 100


    # -------------------------------------------------
    # LOCAL PART ÖNCELİĞİ
    # -------------------------------------------------

    if kullanici in {

        "info",
        "iletisim",
        "contact"

    }:

        puan += 80


    elif kullanici in {

        "satis",
        "sales",
        "pazarlama",
        "muhasebe",
        "accounting"

    }:

        puan += 70


    elif kullanici in {

        "ofis",
        "office",
        "destek",
        "support"

    }:

        puan += 50


    # -------------------------------------------------
    # DOMAIN İSMİ LOCAL PART'TA VAR MI?
    # -------------------------------------------------

    domain_kok = re.sub(
        r"[^a-z0-9]",
        "",
        site_domain
    )

    kullanici_temiz = re.sub(
        r"[^a-z0-9]",
        "",
        kullanici
    )

    if (
        domain_kok
        and domain_kok in kullanici_temiz
    ):

        puan += 120


    # -------------------------------------------------
    # ÜCRETSİZ MAIL
    # -------------------------------------------------

    ucretsiz = {

        "gmail.com",
        "hotmail.com",
        "hotmail.com.tr",
        "outlook.com",
        "outlook.com.tr",
        "yahoo.com",
        "yahoo.com.tr"

    }

    if mail_domain in ucretsiz:

        # yine de kullanılabilir
        puan += 10


    return puan


def en_iyi_mail(mailler, site_domain):

    if not mailler:

        return ""


    puanli = []

    for mail in mailler:

        puan = mail_puani(
            mail,
            site_domain
        )

        puanli.append(
            (
                puan,
                mail
            )
        )


    puanli.sort(
        key=lambda x: x[0],
        reverse=True
    )

    return puanli[0][1]


# =====================================================
# İLÇE BUL
# =====================================================

def ilce_bul(metin):

    metin = temizle(
        metin
    )

    bulunan = []

    for ilce in ISTANBUL_ILCELERI:

        pattern = rf"\b{re.escape(ilce)}\b"

        if re.search(
            pattern,
            metin
        ):

            bulunan.append(
                ilce
            )

    return bulunan


# =====================================================
# ADRES BÖLÜMÜ BUL
# =====================================================

def adres_bolumu_bul(body):

    """
    Sayfanın tamamını adres olarak kabul etmez.

    Adres / İletişim / Contact gibi başlıklardan
    yakınındaki metni almaya çalışır.
    """

    text = str(body)

    satirlar = [
        x.strip()
        for x in text.splitlines()
        if x.strip()
    ]

    anahtarlar = [

        "adres",
        "address",
        "iletişim",
        "iletisim",
        "contact",
        "contact us",
        "bize ulaşın",
        "bize ulasin",
        "office",
        "merkez"

    ]

    aday = []

    for i, satir in enumerate(satirlar):

        satir_temiz = temizle(
            satir
        )

        uygun = False

        for anahtar in anahtarlar:

            if anahtar in satir_temiz:

                uygun = True
                break

        if not uygun:
            continue


        # Başlığın bulunduğu satır + sonraki birkaç satır
        baslangic = max(
            0,
            i
        )

        bitis = min(
            len(satirlar),
            i + 8
        )

        parca = " ".join(
            satirlar[
                baslangic:bitis
            ]
        )

        aday.append(
            parca
        )


    # Eğer başlık bulunamazsa,
    # açık adres patternlerini arayalım

    if not aday:

        adres_patternleri = [

            r".{0,150}mah\.?.{0,150}istanbul",
            r".{0,150}mahallesi.{0,150}istanbul",
            r".{0,150}cad\.?.{0,150}istanbul",
            r".{0,150}sok\.?.{0,150}istanbul",
            r".{0,150}\b\d{5}\b.{0,150}istanbul"

        ]

        temiz_body = " ".join(
            satirlar
        )

        for pattern in adres_patternleri:

            eslesmeler = re.findall(
                pattern,
                temiz_body,
                re.IGNORECASE
            )

            aday.extend(
                eslesmeler
            )


    return "\n".join(
        aday
    )


# =====================================================
# ADRESTEN İLÇE BUL
# =====================================================

def adresten_ilce_bul(adres):

    adres = temizle(
        adres
    )

    # İlçe listesinden kontrol
    bulunan = []

    for ilce in ISTANBUL_ILCELERI:

        pattern = rf"\b{re.escape(ilce)}\b"

        if re.search(
            pattern,
            adres
        ):

            bulunan.append(
                ilce
            )


    if bulunan:

        return bulunan[0]


    # İstanbul kelimesi
    if re.search(
        r"\bistanbul\b",
        adres
    ):

        return "istanbul"


    return ""


# =====================================================
# SAYFA VERİSİ
# =====================================================

def sayfa_verisi_al():

    try:

        body = driver.find_element(
            By.TAG_NAME,
            "body"
        ).text[:MAX_BODY]

    except:

        body = ""


    try:

        title = driver.title

    except:

        title = ""


    meta_description = ""

    try:

        meta = driver.find_element(
            By.XPATH,
            "//meta[@name='description']"
        )

        meta_description = (
            meta.get_attribute(
                "content"
            )
            or ""
        )

    except:

        pass


    metin = (
        title
        + "\n"
        + body
        + "\n"
        + meta_description
    )


    # -------------------------------------------------
    # MAIL
    # -------------------------------------------------

    mailler = mailleri_bul(
        metin
    )


    # -------------------------------------------------
    # ADRES BÖLÜMÜ
    # -------------------------------------------------

    adres = adres_bolumu_bul(
        body
    )


    ilce = adresten_ilce_bul(
        adres
    )


    return {

        "metin": metin,

        "body": body,

        "mail": mailler,

        "adres": adres,

        "ilce": ilce,

        "title": title
    }


# =====================================================
# İLETİŞİM LİNKLERİNİ BUL
# =====================================================

def iletisim_linklerini_bul():

    linkler_sonuc = []

    anahtarlar = [

        "iletisim",
        "contact",
        "adres",
        "address",
        "bize ulasin",
        "bize ulaşın",
        "location",
        "office"

    ]


    try:

        linkler = driver.find_elements(
            By.TAG_NAME,
            "a"
        )


        for link in linkler:

            try:

                href = (
                    link.get_attribute(
                        "href"
                    )
                    or ""
                )

                link_text = temizle(
                    link.text
                )


                # -----------------------------------------
                # MAILTO KESİNLİKLE YOK
                # -----------------------------------------

                if href.lower().startswith(
                    "mailto:"
                ):

                    continue


                if href.lower().startswith(
                    "tel:"
                ):

                    continue


                href_temiz = temizle(
                    href
                )


                uygun = False

                for kelime in anahtarlar:

                    if kelime in link_text:

                        uygun = True
                        break

                    if kelime in href_temiz:

                        uygun = True
                        break


                if uygun:

                    if href not in linkler_sonuc:

                        linkler_sonuc.append(
                            href
                        )


            except:

                continue


    except:

        pass


    return linkler_sonuc


# =====================================================
# SITE DOĞRULA
# =====================================================

def site_dogrula(
    unvan,
    ito_ilce,
    url
):

    try:

        domain = domain_adi(
            url
        )


        # =================================================
        # ANA SAYFA
        # =================================================

        print(
            f"    → Site açılıyor: {url}"
        )


        try:

            driver.set_page_load_timeout(
                10
            )

            driver.get(
                url
            )

        except:

            try:

                driver.execute_script(
                    "window.stop();"
                )

            except:

                pass


        WebDriverWait(
            driver,
            7
        ).until(
            EC.presence_of_element_located(
                (
                    By.TAG_NAME,
                    "body"
                )
            )
        )


        time.sleep(
            0.7
        )


        ana_veri = sayfa_verisi_al()


        bulunan_mailler = list(
            ana_veri["mail"]
        )


        web_ilce = (
            ana_veri["ilce"]
        )


        web_adres = (
            ana_veri["adres"]
        )


        # =================================================
        # İLETİŞİM LİNKLERİ
        # =================================================

        iletisim_linkleri = (
            iletisim_linklerini_bul()
        )


        iletisim_linkleri = (
            iletisim_linkleri[
                :MAX_ILETISIM_SAYFASI
            ]
        )


        # =================================================
        # İLETİŞİM SAYFALARINI TARA
        # =================================================

        for link in iletisim_linkleri:

            try:

                print(
                    f"    → İletişim sayfası: {link}"
                )


                driver.get(
                    link
                )


                WebDriverWait(
                    driver,
                    6
                ).until(
                    EC.presence_of_element_located(
                        (
                            By.TAG_NAME,
                            "body"
                        )
                    )
                )


                time.sleep(
                    0.5
                )


                veri = sayfa_verisi_al()


                # Mail ekle
                for mail in veri["mail"]:

                    if mail not in bulunan_mailler:

                        bulunan_mailler.append(
                            mail
                        )


                # Adres bulunduysa daha güçlü kabul et
                if veri["adres"]:

                    web_adres = veri[
                        "adres"
                    ]


                if veri["ilce"]:

                    web_ilce = veri[
                        "ilce"
                    ]


            except Exception as e:

                print(
                    f"    ! İletişim sayfası okunamadı: {e}"
                )

                continue


        # =================================================
        # SITE KİMLİK PUANI
        # =================================================

        title = ana_veri["title"]

        google_site_puani = google_puani(
            unvan,
            url,
            title
        )


        domain_eslesme = (
            domain_firma_eslesmesi(
                unvan,
                url
            )
        )


        site_puani = (
            domain_eslesme
            + int(
                google_site_puani * 0.35
            )
        )


        # =================================================
        # FİRMA İSMİ SİTEDE GEÇİYOR MU?
        # =================================================

        firma_kelimeleri_liste = (
            firma_kelimeleri(
                unvan
            )
        )


        site_metni = temizle(
            ana_veri["metin"]
        )


        site_kelime_eslesmesi = 0

        for kelime in firma_kelimeleri_liste:

            if kelime in site_metni:

                site_kelime_eslesmesi += 1


        if site_kelime_eslesmesi >= 2:

            site_puani += 35

        elif site_kelime_eslesmesi == 1:

            site_puani += 15


        # =================================================
        # ADRES DURUMU
        # =================================================

        ito_ilce_temiz = temizle(
            ito_ilce
        ).strip()


        web_ilce_temiz = temizle(
            web_ilce
        ).strip()


        if not web_ilce:

            adres_durumu = (
                "ADRES BULUNAMADI"
            )

        elif (
            web_ilce_temiz
            == ito_ilce_temiz
        ):

            adres_durumu = (
                "ADRES AYNI"
            )

        else:

            adres_durumu = (
                "ILCE FARKLI"
            )


        # =================================================
        # MAIL SEÇ
        # =================================================

        secilen_mail = en_iyi_mail(
            bulunan_mailler,
            domain
        )


        # =================================================
        # SONUÇ
        # =================================================

        # Site kimliği yeterince güçlü değilse
        # bu sitenin mailini kesinlikle kullanma.

        if site_puani < MIN_SITE_PUAN:

            return {

                "dogru": False,

                "site_puani":
                    site_puani,

                "mail": "",

                "web_ilce":
                    web_ilce,

                "web_adres":
                    web_adres,

                "adres_durumu":
                    adres_durumu,

                "durum":
                    "SITE ISMI ZAYIF"
            }


        # =================================================
        # ADRES VARSA
        # =================================================

        if web_ilce:

            if (
                web_ilce_temiz
                == ito_ilce_temiz
            ):

                durum = (
                    "DOĞRULANDI"
                )

                dogru = True

            else:

                durum = (
                    "SITE BULUNDU - İLÇE FARKLI"
                )

                # Yüksek kaliteli site ise
                # yine de kabul et.
                dogru = (
                    site_puani
                    >= MIN_SITE_PUAN
                )


        else:

            durum = (
                "SITE BULUNDU - ADRES YOK"
            )

            # Site çok güçlü ise adres bulunmasa
            # bile manuel kontrol için kaydet.
            dogru = (
                site_puani
                >= 100
            )


        return {

            "dogru": dogru,

            "site_puani":
                site_puani,

            "mail":
                secilen_mail,

            "web_ilce":
                web_ilce,

            "web_adres":
                web_adres,

            "adres_durumu":
                adres_durumu,

            "durum":
                durum
        }


    except Exception as e:

        print(
            "    ! Site doğrulama hatası:",
            e
        )

        return {

            "dogru": False,

            "site_puani": 0,

            "mail": "",

            "web_ilce": "",

            "web_adres": "",

            "adres_durumu":
                "SITE HATASI",

            "durum":
                "SITE HATASI"
        }


# =====================================================
# FİRMALARI TARA
# =====================================================

sonuclar = []


for i, (_, row) in enumerate(
    df.iterrows(),
    start=1
):

    firma = str(
        row["UNVAN"]
    ).strip()


    ito_adres = str(
        row["ADRES"]
    ).strip()


    ito_ilce = str(
        row["ILCE"]
    ).strip()


    print(
        "\n"
        + "=" * 70
    )


    print(
        f"[{i}/{len(df)}] {firma}"
    )


    print(
        f"İTO İlçe : {ito_ilce}"
    )


    print(
        f"İTO Adres: {ito_adres}"
    )


    bulunan_site = ""

    bulunan_mail = ""

    bulunan_web_ilce = ""

    bulunan_web_adres = ""

    durum = (
        "SITE BULUNAMADI"
    )

    site_puani = 0

    adres_durumu = ""


    try:

        # =================================================
        # GOOGLE
        # =================================================

        print(
            "  → Google aranıyor..."
        )


        driver.get(
            "https://www.google.com/search?q="
            + quote(firma)
        )


        WebDriverWait(
            driver,
            20
        ).until(
            EC.presence_of_element_located(
                (
                    By.ID,
                    "search"
                )
            )
        )


        time.sleep(
            random.uniform(
                *BEKLEME
            )
        )


        google_sonuclari = (
            driver.find_elements(
                By.CSS_SELECTOR,
                "div.yuRUbf"
            )
        )


        adaylar = []


        # =================================================
        # GOOGLE SONUÇLARI
        # =================================================

        for sira, sonuc in enumerate(
            google_sonuclari[
                :MAX_GOOGLE_SONUC
            ],
            start=1
        ):

            try:

                link = sonuc.find_element(
                    By.TAG_NAME,
                    "a"
                )


                baslik = sonuc.find_element(
                    By.TAG_NAME,
                    "h3"
                ).text


                url = link.get_attribute(
                    "href"
                )


            except:

                continue


            if not url:

                continue


            # =================================================
            # MAILTO / TELEFON YOK
            # =================================================

            if url.lower().startswith(
                "mailto:"
            ):

                continue


            if url.lower().startswith(
                "tel:"
            ):

                continue


            url_lower = url.lower()


            # =================================================
            # IGNORE
            # =================================================

            if any(
                x in url_lower
                for x in IGNORE
            ):

                continue


            if url_lower.endswith(
                ".pdf"
            ):

                continue


            try:

                parsed = urlparse(
                    url
                )

                domain = (
                    parsed.netloc.lower()
                )


            except:

                continue


            # =================================================
            # FİRMA İSMİYLE İLİŞKİ KONTROLÜ
            # =================================================

            puan = google_puani(
                firma,
                url,
                baslik
            )


            # Google sırası bonusu
            if sira == 1:

                puan += 30

            elif sira == 2:

                puan += 20

            elif sira == 3:

                puan += 10


            # =================================================
            # DOMAIN VEYA TITLE FİRMA İSMİYLE HİÇ EŞLEŞMİYORSA ALMA
            # =================================================

            domain_eslesme = (
                domain_firma_eslesmesi(
                    firma,
                    url
                )
            )


            title_temiz = temizle(
                baslik
            )


            firma_kelimeleri_liste = (
                firma_kelimeleri(
                    firma
                )
            )


            title_eslesme = sum(
                1
                for kelime
                in firma_kelimeleri_liste
                if kelime in title_temiz
            )


            if (
                domain_eslesme == 0
                and title_eslesme == 0
            ):

                print(
                    f"  Eleme: {domain} "
                    f"(firma adıyla eşleşme yok)"
                )

                continue


            if puan < MIN_GOOGLE_PUAN:

                continue


            if domain in [
                x["domain"]
                for x in adaylar
            ]:

                continue


            adaylar.append({

                "url":
                    f"{parsed.scheme}://{parsed.netloc}/",

                "domain":
                    domain,

                "puan":
                    puan,

                "sira":
                    sira
            })


        # =================================================
        # SIRALA
        # =================================================

        adaylar.sort(
            key=lambda x: (
                x["puan"],
                -x["sira"]
            ),
            reverse=True
        )


        print(
            f"  → {len(adaylar)} uygun site adayı bulundu."
        )


        # =================================================
        # SITELERI KONTROL ET
        # =================================================

        for aday in adaylar:

            print(
                "\n"
                f"  Aday {aday['sira']}: "
                f"{aday['domain']} "
                f"[Google {aday['puan']}]"
            )


            sonuc = site_dogrula(
                firma,
                ito_ilce,
                aday["url"]
            )


            print(
                f"    Site puanı : "
                f"{sonuc['site_puani']}"
            )


            print(
                f"    Web ilçe   : "
                f"{sonuc['web_ilce'] or '-'}"
            )


            print(
                f"    Adres durumu: "
                f"{sonuc['adres_durumu']}"
            )


            print(
                f"    Mail       : "
                f"{sonuc['mail'] or '-'}"
            )


            print(
                f"    Sonuç      : "
                f"{sonuc['durum']}"
            )


            # =================================================
            # SADECE KABUL EDİLEN SİTENİN MAILİNİ AL
            # =================================================

            if sonuc["dogru"]:

                bulunan_site = (
                    aday["url"]
                )


                bulunan_mail = (
                    sonuc["mail"]
                )


                bulunan_web_ilce = (
                    sonuc["web_ilce"]
                )


                bulunan_web_adres = (
                    sonuc["web_adres"]
                )


                site_puani = (
                    sonuc["site_puani"]
                )


                durum = (
                    sonuc["durum"]
                )


                adres_durumu = (
                    sonuc["adres_durumu"]
                )


                print(
                    "\n"
                    "    ✓ SİTE KABUL EDİLDİ"
                )


                break


            else:

                print(
                    "    ✗ Bu aday elendi."
                )


    except Exception as e:

        print(
            "HATA:",
            e
        )

        durum = (
            "GENEL HATA"
        )


    # =================================================
    # SONUCU EKLE
    # =================================================

    sonuclar.append({

        "UNVAN":
            firma,

        "ITO_ADRES":
            ito_adres,

        "ITO_ILCE":
            ito_ilce,

        "WEB":
            bulunan_site,

        "MAIL":
            bulunan_mail,

        "WEB_ILCE":
            bulunan_web_ilce,

        "ADRES_DURUMU":
            adres_durumu,

        "DURUM":
            durum,

        "SITE_PUANI":
            site_puani
    })


# =====================================================
# CHROME KAPAT
# =====================================================

driver.quit()


# =====================================================
# EXCEL
# =====================================================

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DOSYA = "firmalar_web_mail.xlsx"

sonuc_df = pd.DataFrame(
    sonuclar
)

# Excel'e kaydet
sonuc_df.to_excel(
    DOSYA,
    index=False
)


# =====================================================
# EXCEL RENKLENDİRME
# =====================================================

wb = load_workbook(DOSYA)
ws = wb.active


# Açık kırmızı
acik_kirmizi = PatternFill(
    fill_type="solid",
    fgColor="FCE4D6"
)


# Sütun başlıklarının yerlerini bul
basliklar = {}

for col in range(1, ws.max_column + 1):

    baslik = ws.cell(
        row=1,
        column=col
    ).value

    basliklar[baslik] = col


# DURUM sütununu bul
durum_sutunu = basliklar["DURUM"]


# =====================================================
# SATIRLARI RENKLENDİR
# =====================================================

for row in range(2, ws.max_row + 1):

    durum = ws.cell(
        row=row,
        column=durum_sutunu
    ).value

    if durum in [
        "SITE BULUNDU - İLÇE FARKLI",
        "SITE BULUNDU - ADRES YOK"
    ]:

        # Satırın tamamını açık kırmızı yap
        for col in range(1, ws.max_column + 1):

            ws.cell(
                row=row,
                column=col
            ).fill = acik_kirmizi


# =====================================================
# EXCEL KULLANIM KOLAYLIKLARI
# =====================================================

# İlk satırı sabitle
ws.freeze_panes = "A2"

# Filtreleri aç
ws.auto_filter.ref = ws.dimensions


# Sütun genişlikleri
for column in ws.columns:

    max_length = 0

    column_letter = column[0].column_letter

    for cell in column:

        try:

            length = len(str(cell.value))

            if length > max_length:
                max_length = length

        except:

            pass

    ws.column_dimensions[
        column_letter
    ].width = min(
        max_length + 2,
        50
    )


# Kaydet
wb.save(DOSYA)


# =====================================================
# İSTATİSTİK
# =====================================================

bulunan = (
    sonuc_df["WEB"]
    .fillna("")
    .astype(str)
    .str.strip()
    .ne("")
    .sum()
)


mail_bulunan = (
    sonuc_df["MAIL"]
    .fillna("")
    .astype(str)
    .str.strip()
    .ne("")
    .sum()
)


print(
    "\n"
    + "=" * 70
)


print(
    "İŞLEM TAMAMLANDI."
)


print(
    f"Toplam firma    : {len(sonuc_df)}"
)


print(
    f"Bulunan site    : {bulunan}"
)


print(
    f"Mail bulunan    : {mail_bulunan}"
)


print(
    f"Site bulunamayan: "
    f"{len(sonuc_df) - bulunan}"
)


print(
    "\nDosya: firmalar_web_mail.xlsx"
)